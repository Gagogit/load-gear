"""Normalization service (P2b): parse file using reader_profile rules, produce v1 meter_reads."""

from __future__ import annotations

import io
import logging
import uuid
from datetime import datetime, timezone
from zoneinfo import ZoneInfo

import polars as pl

from load_gear.services.ingest.format_detector import detect_file_type

logger = logging.getLogger(__name__)


class NormalizationError(Exception):
    """Raised when normalization fails."""

    def __init__(self, message: str, *, context: dict | None = None):
        super().__init__(message)
        self.context = context or {}


def normalize(
    raw_bytes: bytes,
    rules: dict,
    *,
    meter_id: str,
    job_id: uuid.UUID,
    source_file_id: uuid.UUID,
) -> tuple[list[dict], dict]:
    """Normalize raw file bytes into v1 meter_read rows using detected rules.

    Returns (rows, quality_stats) where:
      rows: list of dicts ready for bulk_insert
      quality_stats: {total_rows, valid_rows, invalid_rows, warnings}
    """
    header_row = rules.get("header_row", 0)
    timestamp_columns = rules["timestamp_columns"]
    value_column = rules["value_column"]
    date_format = rules["date_format"]
    time_format = rules.get("time_format", "")
    decimal_separator = rules["decimal_separator"]
    unit = rules["unit"]
    series_type = rules["series_type"]
    tz_name = rules.get("timezone", "Europe/Berlin")

    file_type = rules.get("file_type", detect_file_type(raw_bytes))

    if file_type in ("xlsx", "xls"):
        df = _read_excel(raw_bytes, file_type, header_row)
    else:
        df = _read_csv(raw_bytes, rules)

    total_rows = len(df)
    if total_rows == 0:
        raise NormalizationError("File contains no data rows")

    warnings: list[str] = []

    # Build timestamp column
    try:
        df = _build_timestamps(df, timestamp_columns, date_format, time_format, tz_name)
    except Exception as exc:
        ts_samples = []
        try:
            ts_col = timestamp_columns[0]
            if ts_col in df.columns:
                ts_samples = df[ts_col].head(5).to_list()
        except Exception:
            pass
        raise NormalizationError(
            f"Timestamp parsing failed: {exc}",
            context={"column": timestamp_columns, "sample_values": ts_samples},
        ) from exc

    # Parse value column
    try:
        df = _parse_values(df, value_column, decimal_separator)
    except Exception as exc:
        val_samples = []
        try:
            if value_column in df.columns:
                val_samples = df[value_column].head(5).to_list()
        except Exception:
            pass
        raise NormalizationError(
            f"Value parsing failed: {exc}",
            context={"column": value_column, "sample_values": val_samples},
        ) from exc

    # Filter out rows with null timestamps or values
    before_filter = len(df)
    df = df.filter(
        pl.col("ts_utc").is_not_null() & pl.col("parsed_value").is_not_null()
    )
    invalid_rows = before_filter - len(df)
    if invalid_rows > 0:
        warnings.append(f"{invalid_rows} rows dropped due to parse errors")

    valid_rows = len(df)
    if valid_rows == 0:
        raise NormalizationError(
            "Zero valid rows after parsing — check format rules",
            context={"total_rows": total_rows},
        )

    # Convert cumulative to interval if needed
    if series_type == "cumulative":
        df = _cumulative_to_interval(df)
        valid_rows = len(df)
        warnings.append("Converted cumulative values to interval deltas")

    # Normalize unit (Wh → kWh, MWh → kWh)
    target_unit = unit
    if unit == "Wh":
        df = df.with_columns(pl.col("parsed_value") / 1000.0)
        target_unit = "kWh"
        warnings.append("Converted Wh to kWh")
    elif unit == "MWh":
        df = df.with_columns(pl.col("parsed_value") * 1000.0)
        target_unit = "kWh"
        warnings.append("Converted MWh to kWh")

    # Build output rows
    rows: list[dict] = []
    for row in df.iter_rows(named=True):
        rows.append({
            "ts_utc": row["ts_utc"],
            "meter_id": meter_id,
            "version": 1,
            "job_id": job_id,
            "value": round(row["parsed_value"], 4),
            "unit": target_unit,
            "quality_flag": 0,
            "source_file_id": source_file_id,
        })

    quality_stats = {
        "total_rows": total_rows,
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "warnings": warnings,
    }

    return rows, quality_stats


def _read_csv(raw_bytes: bytes, rules: dict) -> pl.DataFrame:
    """Read CSV bytes into a Polars DataFrame."""
    encoding = rules["encoding"]
    delimiter = rules["delimiter"]
    header_row = rules.get("header_row", 0)

    try:
        text = raw_bytes.decode(encoding)
    except (UnicodeDecodeError, LookupError) as exc:
        raise NormalizationError(f"Cannot decode file with encoding {encoding}: {exc}") from exc

    if text.startswith("\ufeff"):
        text = text[1:]

    lines = text.split("\n")
    if header_row > 0:
        text = "\n".join(lines[header_row:])

    try:
        return pl.read_csv(
            io.StringIO(text),
            separator=delimiter,
            has_header=True,
            infer_schema_length=0,
        )
    except Exception as exc:
        raise NormalizationError(f"Polars CSV parsing failed: {exc}") from exc


def _read_excel(raw_bytes: bytes, file_type: str, header_row: int) -> pl.DataFrame:
    """Read XLS/XLSX bytes into a Polars DataFrame (all columns as strings)."""
    if file_type == "xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        ws = wb.active
        all_rows: list[list[str]] = []
        for row in ws.iter_rows():
            all_rows.append(
                [str(cell.value) if cell.value is not None else "" for cell in row]
            )
        wb.close()
    else:
        import xlrd

        wb = xlrd.open_workbook(file_contents=raw_bytes)
        ws = wb.sheet_by_index(0)
        all_rows = []
        for rx in range(ws.nrows):
            all_rows.append([str(ws.cell_value(rx, cx)) for cx in range(ws.ncols)])

    if len(all_rows) <= header_row:
        raise NormalizationError("Excel file has fewer rows than expected header_row")

    columns = [c.strip() for c in all_rows[header_row]]
    data_rows = all_rows[header_row + 1:]

    if not data_rows:
        raise NormalizationError("File contains no data rows")

    # Build DataFrame from rows
    data_dict: dict[str, list[str]] = {col: [] for col in columns}
    for row in data_rows:
        for i, col in enumerate(columns):
            val = row[i] if i < len(row) else ""
            data_dict[col].append(val)

    return pl.DataFrame(data_dict)


def _build_timestamps(
    df: pl.DataFrame,
    timestamp_columns: list[str],
    date_format: str,
    time_format: str,
    tz_name: str,
) -> pl.DataFrame:
    """Parse and combine timestamp columns, convert to UTC."""
    tz = ZoneInfo(tz_name)

    if len(timestamp_columns) == 1 and (not time_format or time_format == ""):
        # Combined datetime column
        col_name = timestamp_columns[0]
        fmt = date_format  # already a combined format like %Y-%m-%d %H:%M
        df = df.with_columns(
            pl.col(col_name).str.strip_chars().str.strptime(
                pl.Datetime, fmt, strict=False
            ).alias("ts_local")
        )
    elif len(timestamp_columns) == 2:
        # Separate date + time columns
        date_col, time_col = timestamp_columns[0], timestamp_columns[1]
        combined_fmt = f"{date_format} {time_format}"
        df = df.with_columns(
            (pl.col(date_col).str.strip_chars() + pl.lit(" ") + pl.col(time_col).str.strip_chars())
            .str.strptime(pl.Datetime, combined_fmt, strict=False)
            .alias("ts_local")
        )
    elif len(timestamp_columns) == 1 and time_format:
        # Single date column + time in the format
        col_name = timestamp_columns[0]
        combined_fmt = f"{date_format} {time_format}"
        df = df.with_columns(
            pl.col(col_name).str.strip_chars().str.strptime(
                pl.Datetime, combined_fmt, strict=False
            ).alias("ts_local")
        )
    else:
        raise NormalizationError(
            f"Unsupported timestamp column configuration: {timestamp_columns}"
        )

    # Convert local time to UTC using Python datetime (handles DST correctly)
    #
    # DST fall-back disambiguation: during the ambiguous hour (e.g. 2024-10-27
    # 02:00-02:59 in Europe/Berlin), fold=0 → CEST (first occurrence),
    # fold=1 → CET (second occurrence).  We track seen UTC timestamps and
    # flip to fold=1 on collision so both occurrences get distinct UTC values.
    ts_utc_values: list[datetime | None] = []
    seen_utc: set[datetime] = set()
    for ts_local in df["ts_local"].to_list():
        if ts_local is None:
            ts_utc_values.append(None)
            continue
        try:
            local_dt = ts_local.replace(tzinfo=tz, fold=0)
            utc_dt = local_dt.astimezone(timezone.utc)
            if utc_dt in seen_utc:
                # DST fall-back collision: try the second occurrence (fold=1)
                local_dt = ts_local.replace(tzinfo=tz, fold=1)
                utc_dt = local_dt.astimezone(timezone.utc)
            seen_utc.add(utc_dt)
            ts_utc_values.append(utc_dt)
        except Exception:
            ts_utc_values.append(None)

    df = df.with_columns(
        pl.Series("ts_utc", ts_utc_values, dtype=pl.Datetime("us", "UTC"))
    )

    return df


def _parse_values(
    df: pl.DataFrame, value_column: str, decimal_separator: str
) -> pl.DataFrame:
    """Parse the value column into floats."""
    if decimal_separator == ",":
        df = df.with_columns(
            pl.col(value_column)
            .str.strip_chars()
            .str.replace_all(r"\.", "")  # remove thousands dots
            .str.replace(",", ".")
            .cast(pl.Float64, strict=False)
            .alias("parsed_value")
        )
    else:
        df = df.with_columns(
            pl.col(value_column)
            .str.strip_chars()
            .str.replace_all(",", "")  # remove thousands commas
            .cast(pl.Float64, strict=False)
            .alias("parsed_value")
        )
    return df


def _cumulative_to_interval(df: pl.DataFrame) -> pl.DataFrame:
    """Convert cumulative meter readings to interval deltas."""
    df = df.sort("ts_utc")
    df = df.with_columns(
        (pl.col("parsed_value") - pl.col("parsed_value").shift(1)).alias("parsed_value_delta")
    )
    # Drop first row (no delta) and use delta as the value
    df = df.slice(1)
    df = df.with_columns(pl.col("parsed_value_delta").alias("parsed_value"))
    df = df.drop("parsed_value_delta")
    return df
